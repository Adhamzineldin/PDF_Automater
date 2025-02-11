import os
import sys
import subprocess
import tempfile
from collections.abc import Iterable

from svgpathtools import svg2paths
from PIL import Image, ImageDraw

from ACCAPI import ACCAPI

# Decide which backend to use based on the OS.
USE_XLWINGS = sys.platform.startswith('win')

if USE_XLWINGS:
    import xlwings as xw
else:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.properties import PageSetupProperties


class ExcelModifier:
    def __init__(self, template_filename, modified_folder):
        self.script_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_path = os.path.join(self.script_dir, template_filename)
        self.modified_folder = os.path.join(self.script_dir, modified_folder)

        if not os.path.exists(self.modified_folder):
            os.makedirs(self.modified_folder)

        self.backend = 'xlwings' if USE_XLWINGS else 'openpyxl'
        self.app = None
        self.workbook = None
        self.sheet = None

    def open_workbook(self):
        """Opens the Excel workbook and initializes the sheet."""
        if self.backend == 'xlwings':
            self.app = xw.App(visible=False, add_book=False)
            self.workbook = self.app.books.open(self.excel_path)
            self.sheet = self.workbook.sheets[0]
        else:
            self.workbook = openpyxl.load_workbook(self.excel_path)
            self.sheet = self.workbook.active
        print(f"Workbook opened using {self.backend}.")

    def modify_cell(self, cell_range, value):
        """Modifies a specific cell range with a new value."""
        if self.sheet is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")
        if self.backend == 'xlwings':
            cell = self.sheet.range(cell_range)
            cell.value = value
        else:
            # For openpyxl, handle merged cells properly
            cell = self.sheet[cell_range.split(":")[0]]  # Get the top-left cell of the range
    
            # Check if the cell is part of a merged range using openpyxl's merged_cells method
            is_merged = False
            cell_coord = cell.coordinate  # Get the coordinate of the cell (like 'A1')
    
            for merged_range in self.sheet.merged_cells.ranges:
                if cell_coord in merged_range:
                    is_merged = True
                    break
    
            if is_merged:
                # If the cell is merged, modify the top-left cell of the merged range
                for merged_range in self.sheet.merged_cells.ranges:
                    if cell_coord in merged_range:
                        top_left_cell = self.sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = value
                        break
            else:
                # Modify the original cell if it's not part of a merged range
                cell.value = value
        print(f"Cell {cell_range} updated to {value}.")

    def auto_fit_columns(self):
        """Automatically adjusts all columns to fit content."""
        if self.sheet is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")

        if self.backend == 'xlwings':
            self.sheet.api.Columns.AutoFit()
        else:
            # For openpyxl, iterate through all columns in the sheet and adjust width
            for col in self.sheet.columns:
                max_length = 0
                column = col[0].column  # can be a number or letter
                col_letter = get_column_letter(column) if isinstance(column, int) else column
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                # Add a little extra space
                adjusted_width = (max_length + 2)
                self.sheet.column_dimensions[col_letter].width = adjusted_width
        print("Auto-fit applied to all columns.")

    def add_gridlines(self):
        """
        Adds gridlines to the sheet by applying borders.
        Note: On Windows, Excel shows gridlines by default.
        """
        if self.sheet is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")

        if self.backend == 'xlwings':
            # Use borders to simulate gridlines (as in the original code)
            for border_id in range(7, 13):  # Border IDs for Excel
                self.sheet.api.Cells.Borders(border_id).LineStyle = 1  # xlContinuous
        else:
            # For openpyxl, we add a thin border to each cell in the used range.
            from openpyxl.styles import Border, Side
            thin = Side(style='thin', color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Determine the used range (assuming sheet.max_row and sheet.max_column)
            for row in self.sheet.iter_rows(min_row=1, max_row=self.sheet.max_row,
                                            min_col=1, max_col=self.sheet.max_column):
                for cell in row:
                    cell.border = border
        print("Gridlines added to the sheet.")

    # def insert_row(self, row):
    #     """Inserts a new row using xlwings by shifting down rows manually."""
    #     self.sheet.range(f"{row}:{row}").insert(shift="down")
    #     print(f"Inserted a new row at {row}.")


    def insert_row(self, row):
        """Inserts a new row and copies the styling from the row above."""
        self.sheet.insert_rows(row)
    
        # Copy styling from the row above
        if row > 1:
            for col in range(1, self.sheet.max_column + 1):
                cell_above = self.sheet.cell(row=row - 1, column=col)
                new_cell = self.sheet.cell(row=row, column=col)
    
                # Copy style attributes
                if cell_above.has_style:
                    new_cell.font = cell_above.font
                    new_cell.fill = cell_above.fill
                    new_cell.border = cell_above.border
                    new_cell.alignment = cell_above.alignment
                    new_cell.number_format = cell_above.number_format
    
        print(f"Inserted a new row at {row} with styling copied from the row above.")


    def save_workbook(self, filename='modified.xlsx'):
        """Saves the workbook with a new name."""
        if self.workbook is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")
        save_path = os.path.join(self.modified_folder, filename)
        if self.backend == 'xlwings':
            self.workbook.save(save_path)
        else:
            self.workbook.save(save_path)
        print(f"Workbook saved at {save_path}")
        return save_path

    def export_to_pdf(self, payment=None, filename='modified.pdf', excel_filename="output", project_name="Information Systems Workspace", destination_folder="Cost Cover Sheets"):
        """Exports the sheet to a PDF, fitting it to a single page."""
        if self.sheet is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")


        name = None
        if payment["number"]:
            name = f'{payment["number"]}_{payment["status"]}'
            print(name)
        else:
            name = f"{excel_filename}"
            print(name)
    
        pdf_path = os.path.join(self.modified_folder, name)
        print(pdf_path)
    
        if self.backend == 'xlwings':
            # Windows-specific export using xlwings (unchanged)
            sheet_api = self.sheet.api
            sheet_api.PageSetup.FitToPagesWide = 1  # Fit to one page wide
            sheet_api.PageSetup.FitToPagesTall = 1   # Fit to one page tall
            sheet_api.PageSetup.Zoom = False         # Disable zoom
            try:
                sheet_api.ExportAsFixedFormat(0, pdf_path)  # 0 refers to xlTypePDF
                print(f"PDF exported at {pdf_path}")
            except Exception as e:
                print(f"Error exporting to PDF: {e}")
                return None
        else:
            # For Linux, use LibreOffice in headless mode to convert the saved XLSX to PDF.
            temp_xlsx = f"modified_files/{excel_filename}.xlsx"
            # Capture the current working directory
            original_dir = os.getcwd()
            
            try:
                cmd = [
                        'libreoffice', '--headless',
                        '--convert-to', 'pdf',
                        '--outdir', self.modified_folder,
                        temp_xlsx
                ]



                subprocess.run(cmd, check=True)
    
                # Ensure that the generated PDF has the same name as the input XLSX file.
                generated_pdf = os.path.join(self.modified_folder, f'{excel_filename}.pdf')
                print(generated_pdf)
    
                # If the output file already exists, delete it to avoid conflicts.
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
    
                # Rename the generated PDF to the desired filename (overwrite if exists).
                os.rename(generated_pdf, pdf_path)
                print(f"PDF exported at {pdf_path}")


                acc_api = ACCAPI()
                
                
            
                acc_api.upload_pdf_to_acc(pdf_path=pdf_path, filename=name, project_name=project_name, folder_name=destination_folder)
                
                
                
                
            except subprocess.CalledProcessError as e:
                print(f"Error exporting to PDF via LibreOffice: {e}")
                return None
    
        return pdf_path


    def export_to_pdf_no_upload(self, excel_filename="output"):
        """Exports the sheet to a PDF, fitting it to a single page."""
        if self.sheet is None:
            raise Exception("Workbook is not opened. Call open_workbook() first.")
    
        # Construct paths and ensure directory exists
        modified_folder_path = "modified_files"
        os.makedirs(modified_folder_path, exist_ok=True)
    
        temp_xlsx = os.path.join(modified_folder_path, f"{excel_filename}.xlsx")
        if not os.path.exists(temp_xlsx):
            print(f"Error: {temp_xlsx} does not exist.")
            return None
    
        pdf_path = os.path.join(modified_folder_path, f"{excel_filename}.pdf")
    
        # Convert using LibreOffice
        try:
            cmd = [
                    'libreoffice', '--headless',
                    '--convert-to', 'pdf',
                    '--outdir', self.modified_folder,
                    temp_xlsx  # No quotes needed; subprocess handles spaces
            ]
            result = subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print(f"LibreOffice stdout: {result.stdout.decode()}")
            print(f"LibreOffice stderr: {result.stderr.decode()}")
    
            # Verify PDF was created
            if not os.path.exists(pdf_path):
                print(f"Error: PDF not generated at {pdf_path}")
                return None
    
            print(f"PDF exported at {pdf_path}")
            return pdf_path
    
        except subprocess.CalledProcessError as e:
            print(f"LibreOffice conversion failed: {e.stderr.decode()}")
            return None


    def insert_svg_as_image(self, svg_code, cell_range):
        """
        Converts SVG code to PNG using svgpathtools and Pillow, and inserts it into the Excel sheet.

        Parameters:
        - svg_code: str, SVG code as a string.
        - cell_range: str, Excel cell range where the image should be inserted.
        """
        try:
            # Step 1: Save the SVG code to a temporary file
            temp_svg_path = os.path.join(self.modified_folder, "temp_image.svg")
            with open(temp_svg_path, "w", encoding="utf-8") as svg_file:
                svg_file.write(svg_code)

            # Step 2: Parse the SVG to extract paths
            paths, attributes = svg2paths(temp_svg_path)

            # Step 3: Create a new blank image (transparent background)
            width, height = 600, 300  # adjust size as needed
            img = Image.new('RGBA', (width, height), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)

            # Step 4: Draw the paths onto the image
            for path in paths:
                for segment in path:
                    start = segment.start
                    end = segment.end
                    draw.line((start.real, start.imag, end.real, end.imag), fill='black', width=2)

            # Step 5: Save the image as PNG
            temp_png_path = os.path.join(self.modified_folder, "temp_image.png")
            img.save(temp_png_path, "PNG")

            # Step 6: Insert the PNG into the Excel sheet
            if self.backend == 'xlwings':
                self.sheet.pictures.add(temp_png_path,
                                        left=self.sheet.range(cell_range).left,
                                        top=self.sheet.range(cell_range).top)
            else:
                # For openpyxl, create an image object and anchor it to the cell.
                xl_img = XLImage(temp_png_path)
                xl_img.anchor = cell_range  # e.g. "B2"
                self.sheet.add_image(xl_img)
            print(f"SVG inserted as image at {cell_range}")
        except Exception as e:
            print(f"An error occurred while processing the SVG: {e}")

    def close_workbook(self):
        """Closes the workbook and Excel application if necessary."""
        if self.backend == 'xlwings':
            if self.workbook:
                self.workbook.close()
            if self.app:
                self.app.quit()
        # For openpyxl, nothing special is needed.
        print("Workbook closed.")


# === Example usage ===
if __name__ == "__main__":
    # Update these file names and paths as needed.
    template_filename = "template.xlsx"
    modified_folder = "modified_files"

    modifier = ExcelModifier(template_filename, modified_folder)
    modifier.open_workbook()
    modifier.modify_cell("B2", "Hello, World!")
    modifier.auto_fit_columns()
    modifier.add_gridlines()

    # Sample SVG code (a simple line for demonstration)
    sample_svg = '''<svg height="100" width="100" xmlns="http://www.w3.org/2000/svg">
                      <line x1="0" y1="0" x2="100" y2="100" stroke="black" stroke-width="2"/>
                    </svg>'''
    modifier.insert_svg_as_image(sample_svg, "D4")

    # Save workbook and export to PDF
    modifier.save_workbook("modified.xlsx")
    modifier.export_to_pdf("modified.pdf")
    modifier.close_workbook()
